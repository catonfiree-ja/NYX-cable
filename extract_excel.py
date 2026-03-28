import openpyxl
import json

wb = openpyxl.load_workbook(r'C:\Users\ปิ๊ก\Downloads\SpecNYXcable.xlsx')

with open('excel_data_utf8.txt', 'w', encoding='utf-8') as f:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write(f"\n{'='*80}\n")
        f.write(f"=== Sheet: {sheet_name} (Rows: {ws.max_row}, Cols: {ws.max_column}) ===\n")
        f.write(f"{'='*80}\n")
        for row in range(1, ws.max_row + 1):
            vals = []
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    vals.append(f"C{col}:{v}")
            if vals:
                f.write(f"  R{row}: {' | '.join(vals)}\n")
    
print("Done writing excel_data_utf8.txt")
