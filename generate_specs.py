"""
Generate product-specs.json from SpecNYXcable.xlsx
Reads all 9 sheets and produces a JSON file matching the website's expected format.
"""
import json
import openpyxl
import os
import re

EXCEL_PATH = r"C:\Users\ปิ๊ก\Downloads\SpecNYXcable.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), "frontend", "data", "product-specs.json")

# Sheet configurations: maps sheet name -> product slug(s) and URL pattern
# Each sheet in Excel has data rows starting at row 6, with columns:
#   C2=Part NO, C8=Core x Size, C9=Model, C10=strands, C11=Outer Dia,
#   C12=Cu Weight, C13=Weight, C14=Resistance, C15=Price, C17=Link

SHEET_CONFIG = {
    "YSLY": {
        "products": {
            # Merge OZ + JZ into single ysly-jz product (no Sanity CMS entry for ysly-oz)
            "ysly-jz": {"name_prefix": "สายคอนโทรล YSLY-JZ", "filter_slug": None},
        }
    },
    "JZ-500": {
        "products": {
            "jz-500": {"name_prefix": None}  # Uses C9 Model directly
        }
    },
    "Olflex": {
        "products": {
            "olflex-classic-110": {"name_prefix": None}
        }
    },
    "Flex-JZ": {
        "products": {
            "flex-jz": {"name_prefix": None}
        }
    },
    "OPVC-JZ": {
        "products": {
            "opvc-jz": {"name_prefix": None}
        }
    },
    "115": {
        "products": {
            "olflex-classic-115-cy": {"name_prefix": None}
        }
    },
    "LiYCY": {
        "products": {
            "liycy": {"name_prefix": None}
        }
    },
    "LiYCY-JZ": {
        "products": {
            "liycy-jz": {"name_prefix": None}
        }
    },
    "H07V-K": {
        "products": {
            "h07v-k": {"name_prefix": None}
        }
    },
}

def clean_val(v):
    """Convert cell value to clean string or None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "None"):
        return None
    # Check if it's a formula that wasn't computed
    if s.startswith("="):
        return None
    return s

def format_price(v):
    """Format price: round to 2 decimal places, or None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "None") or s.startswith("="):
        return None
    try:
        num = float(s)
        # Format to 2 decimal places, removing trailing zeros but keeping at least one
        formatted = f"{num:.2f}"
        return formatted
    except ValueError:
        return None

def format_number(v):
    """Format numbers: remove .0 from integers, keep decimals as-is."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "None") or s.startswith("="):
        return None
    try:
        num = float(s)
        if num == int(num):
            return str(int(num))
        # Keep the original decimal representation
        # Remove trailing zeros
        formatted = f"{num:g}"
        return formatted
    except ValueError:
        return s

def extract_sheet_data(ws, sheet_name):
    """Extract all data rows from a worksheet."""
    items = []
    
    # Determine Part NO column:
    # - YSLY sheet: Part NO in col3 (col2 is empty)
    # - All other sheets: Part NO in col2
    # Check by looking at the header row (row 4)
    col2_header = ws.cell(row=4, column=2).value
    if col2_header and "Part" in str(col2_header):
        part_no_col = 2  # Non-YSLY sheets
    else:
        part_no_col = 3  # YSLY sheet
    
    # Data starts at row 6
    data_start = 6
    
    for row_num in range(data_start, ws.max_row + 1):
        core_size = ws.cell(row=row_num, column=8).value  # C8
        model = ws.cell(row=row_num, column=9).value       # C9
        
        # Skip rows without core_size and model
        if core_size is None and model is None:
            continue
        
        core_size_str = str(core_size).strip() if core_size else ""
        if core_size_str in ("Core x Conductor", "Size (mm2)", "", "None"):
            continue
            
        # Read Part NO from the correct column
        part_no = ws.cell(row=row_num, column=part_no_col).value
        
        # For YSLY sheet: col4=product type, col5=slug
        # For other sheets: col3=YSLY ref type, col4=YSLY ref slug
        # We use _col5 for YSLY filtering (slug in col5 for YSLY)
        if part_no_col == 3:
            # YSLY sheet
            col5 = ws.cell(row=row_num, column=5).value  # slug for filtering
        else:
            # Non-YSLY: no slug-based filtering needed
            col5 = None
        
        strands = ws.cell(row=row_num, column=10).value
        outer_dia = ws.cell(row=row_num, column=11).value
        cu_weight = ws.cell(row=row_num, column=12).value
        weight = ws.cell(row=row_num, column=13).value
        resistance = ws.cell(row=row_num, column=14).value
        price = ws.cell(row=row_num, column=15).value
        link = ws.cell(row=row_num, column=17).value
        
        model_str = str(model).strip() if model else ""
        if not model_str or model_str == "None":
            continue
        
        item = {
            "partNo": clean_val(part_no),
            "coreSize": clean_val(core_size),
            "model": clean_val(model),
            "strands": clean_val(strands),
            "outerDia": format_number(outer_dia),
            "cuWeight": format_number(cu_weight),
            "weight": format_number(weight),
            "resistance": format_number(resistance),
            "price": format_price(price),
            "link": clean_val(link),
            # Extra metadata for filtering
            "_col5": clean_val(col5),
        }
        items.append(item)
    
    return items

def build_product_data(all_items, sheet_name, product_slug, product_config):
    """Build product data for a specific product slug from sheet items."""
    
    # For YSLY sheet, we need to split by col5 slug (ysly-oz vs ysly-jz)
    if sheet_name == "YSLY":
        filter_slug = product_config.get("filter_slug")
        if filter_slug:
            filtered = [item for item in all_items if item.get("_col5") == filter_slug]
        else:
            filtered = all_items
    else:
        filtered = all_items
    
    # Build final items (remove internal metadata)
    items = []
    for item in filtered:
        clean_item = {
            "partNo": item["partNo"],
            "coreSize": item["coreSize"],
            "model": item["model"],
            "strands": item["strands"],
            "outerDia": item["outerDia"],
            "cuWeight": item["cuWeight"],
            "weight": item["weight"],
            "resistance": item["resistance"],
            "price": item["price"],
            "link": item["link"],
        }
        items.append(clean_item)
    
    # Build product URL from the first item's link or use a default pattern
    product_url = f"/products/detail/{product_slug}"
    
    # Determine sheet display name
    if product_slug == "ysly-oz":
        sheet_display = "สายคอนโทรล YSLY-OZ"
    elif product_slug == "ysly-jz":
        sheet_display = "สายคอนโทรล YSLY-JZ"
    elif product_slug == "jz-500":
        sheet_display = "JZ 500 Volt"
    elif product_slug == "olflex-classic-110":
        sheet_display = "Olflex Classic 110"
    elif product_slug == "flex-jz":
        sheet_display = "Flex-JZ"
    elif product_slug == "opvc-jz":
        sheet_display = "OPVC-JZ"
    elif product_slug == "olflex-classic-115-cy":
        sheet_display = "Olflex Classic 115 CY"
    elif product_slug == "liycy":
        sheet_display = "LiYCY"
    elif product_slug == "liycy-jz":
        sheet_display = "สายชีลด์ LiYCY-JZ"
    elif product_slug == "h07v-k":
        sheet_display = "H07V-K"
    else:
        sheet_display = product_slug
    
    return {
        "sheetName": sheet_display,
        "productUrl": product_url,
        "items": items,
        "count": len(items),
    }

def main():
    print(f"Reading Excel: {EXCEL_PATH}")
    
    # Load with data_only=True to get computed formula values
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    print(f"Available sheets: {wb.sheetnames}")
    
    result = {}
    
    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARNING: Sheet '{sheet_name}' not found, skipping")
            continue
            
        ws = wb[sheet_name]
        print(f"\n  Processing sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)")
        
        # Extract all data rows from sheet
        all_items = extract_sheet_data(ws, sheet_name)
        print(f"    Extracted {len(all_items)} data rows")
        
        # Build products from sheet data
        for product_slug, product_config in config["products"].items():
            product_data = build_product_data(all_items, sheet_name, product_slug, product_config)
            result[product_slug] = product_data
            print(f"    -> {product_slug}: {product_data['count']} items")
    
    # Write output
    print(f"\nWriting to: {OUTPUT_PATH}")
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    # Summary
    print("\n=== Summary ===")
    total = 0
    for slug, data in result.items():
        print(f"  {slug}: {data['count']} items (sheetName: {data['sheetName']})")
        total += data['count']
    print(f"  TOTAL: {total} items across {len(result)} products")

if __name__ == "__main__":
    main()
