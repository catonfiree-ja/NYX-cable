import glob, os

try:
    from openpyxl import load_workbook
except:
    os.system('pip install openpyxl')
    from openpyxl import load_workbook

# Search recursively for xlsx in parent
parent = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))
files = []
for item in os.listdir(parent):
    if item.endswith('.xlsx'):
        files.append(os.path.join(parent, item))

out = 'C:\\tmp_excel.txt'
with open(out, 'w', encoding='utf-8') as f:
    if not files:
        f.write("NO XLSX FOUND\n")
        for item in os.listdir(parent):
            f.write(f"  {item}\n")
    else:
        for fpath in files:
            wb = load_workbook(fpath, data_only=True)
            f.write(f"SHEETS: {wb.sheetnames}\n\n")
            for sname in wb.sheetnames:
                ws = wb[sname]
                f.write(f"--- {sname} ({ws.max_row}r x {ws.max_column}c) ---\n")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 6:
                        f.write("...\n")
                        break
                    f.write(f"{list(row)}\n")
                f.write("\n")
print("DONE")
