import glob
import os

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    from openpyxl import load_workbook

# Find xlsx files in the nyxcable-website root
files = glob.glob(os.path.join(os.path.dirname(__file__), '..', '*.xlsx'))

out = os.path.join(os.path.dirname(__file__), 'excel_output.txt')

with open(out, 'w', encoding='utf-8') as f:
    if not files:
        f.write("NO XLSX FILES FOUND\n")
        f.write("Searched in: " + os.path.abspath(os.path.join(os.path.dirname(__file__), '..')) + "\n")
        # List all files in parent
        parent = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        for item in os.listdir(parent):
            f.write(f"  {item}\n")
    else:
        for fpath in files:
            f.write(f"=== FILE: {os.path.basename(fpath)} ===\n\n")
            wb = load_workbook(fpath, data_only=True)
            f.write(f"SHEETS: {wb.sheetnames}\n\n")
            for sname in wb.sheetnames:
                ws = wb[sname]
                f.write(f"--- Sheet: {sname} ({ws.max_row} rows x {ws.max_column} cols) ---\n")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 8:  # First 8 rows only
                        f.write("  ...(truncated)\n")
                        break
                    f.write(f"  Row {i+1}: {list(row)}\n")
                f.write("\n")

print(f"Done! Output at: {out}")
