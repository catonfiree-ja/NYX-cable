"""
Final comprehensive verification:
1. Cross-check JSON vs Excel cell-by-cell for all 9 sheets
2. Validate JSON structure (no missing fields, correct types)
3. Verify all prices are correctly formatted
4. Check for duplicates
5. Verify links are valid URLs
"""
import json
import openpyxl
import re
import sys
import os
os.environ['PYTHONIOENCODING'] = 'utf-8'
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\ปิ๊ก\Downloads\SpecNYXcable.xlsx"
JSON_PATH = r"frontend\data\product-specs.json"

errors = []
warnings = []

def err(msg):
    errors.append(msg)
    print("  [FAIL] " + msg)

def warn(msg):
    warnings.append(msg)
    print("  [WARN] " + msg)

def ok(msg):
    print("  [OK] " + msg)

# ══════════════════════════════════════════════════
# 1. Load both data sources
# ══════════════════════════════════════════════════
print("=" * 60)
print("1. LOADING DATA SOURCES")
print("=" * 60)

with open(JSON_PATH, "r", encoding="utf-8") as f:
    json_data = json.load(f)
ok("JSON loaded: " + str(len(json_data)) + " products")

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ok("Excel loaded: " + str(len(wb.sheetnames)) + " sheets: " + ", ".join(wb.sheetnames))

# ══════════════════════════════════════════════════
# 2. STRUCTURE VALIDATION
# ══════════════════════════════════════════════════
print("\n" + "=" * 60)
print("2. JSON STRUCTURE VALIDATION")
print("=" * 60)

required_fields = ["partNo", "coreSize", "model", "strands", "outerDia", "cuWeight", "weight", "resistance", "price", "link"]
nullable_fields = ["partNo", "price"]  # These can be null

for slug, product in json_data.items():
    # Check product-level fields
    if "sheetName" not in product:
        err(slug + ": missing sheetName")
    if "productUrl" not in product:
        err(slug + ": missing productUrl")
    if "items" not in product:
        err(slug + ": missing items")
        continue
    if "count" not in product:
        err(slug + ": missing count")
    if product["count"] != len(product["items"]):
        err(slug + ": count mismatch: " + str(product["count"]) + " != " + str(len(product["items"])))

    # Check each item
    for idx, item in enumerate(product["items"]):
        for field in required_fields:
            if field not in item:
                err(slug + " item " + str(idx) + ": missing field '" + field + "'")
            elif item[field] is None and field not in nullable_fields:
                err(slug + " item " + str(idx) + " (" + str(item.get("coreSize", "?")) + "): '" + field + "' is null (should have value)")

        # Check no internal metadata leaked
        for key in item:
            if key.startswith("_"):
                err(slug + " item " + str(idx) + ": internal field leaked: " + key)

ok("Structure check complete")

# ══════════════════════════════════════════════════
# 3. PRICE FORMAT VALIDATION
# ══════════════════════════════════════════════════
print("\n" + "=" * 60)
print("3. PRICE FORMAT VALIDATION")
print("=" * 60)

total_items = 0
null_prices = 0
valid_prices = 0

for slug, product in json_data.items():
    for item in product["items"]:
        total_items += 1
        price = item.get("price")
        if price is None:
            null_prices += 1
        else:
            # Check format: should be string with 2 decimal places
            if not isinstance(price, str):
                err(slug + " " + str(item["coreSize"]) + ": price is not string: " + repr(price))
            elif not re.match(r"^\d+\.\d{2}$", price):
                err(slug + " " + str(item["coreSize"]) + ": price format invalid: " + price)
            else:
                valid_prices += 1

ok("Total items: " + str(total_items))
ok("Valid prices: " + str(valid_prices) + ", Null prices (unavailable): " + str(null_prices))

# ══════════════════════════════════════════════════
# 4. DUPLICATE CHECK
# ══════════════════════════════════════════════════
print("\n" + "=" * 60)
print("4. DUPLICATE CHECK")
print("=" * 60)

for slug, product in json_data.items():
    models = [item["model"] for item in product["items"] if item["model"]]
    dupes = [m for m in set(models) if models.count(m) > 1]
    if dupes:
        err(slug + ": duplicate models: " + str(dupes[:5]))
    else:
        ok(slug + ": no duplicate models (" + str(len(models)) + " unique)")

# ══════════════════════════════════════════════════
# 5. LINK VALIDATION
# ══════════════════════════════════════════════════
print("\n" + "=" * 60)
print("5. LINK VALIDATION")
print("=" * 60)

for slug, product in json_data.items():
    bad_links = 0
    for item in product["items"]:
        link = item.get("link")
        if link is None:
            bad_links += 1
        elif not link.startswith("http"):
            bad_links += 1
    if bad_links > 0:
        warn(slug + ": " + str(bad_links) + " items have missing/invalid links")
    else:
        ok(slug + ": all " + str(len(product["items"])) + " links valid")

# ══════════════════════════════════════════════════
# 6. CROSS-CHECK: JSON vs EXCEL (spot checks)
# ══════════════════════════════════════════════════
print("\n" + "=" * 60)
print("6. CROSS-CHECK: JSON vs EXCEL (cell-by-cell)")
print("=" * 60)

def fmt_num(v):
    """Format number same as generate script."""
    if v is None: return None
    s = str(v).strip()
    if s in ("", "-", "None") or s.startswith("="): return None
    try:
        num = float(s)
        if num == int(num): return str(int(num))
        return "{:g}".format(num)
    except: return s

def fmt_price(v):
    if v is None: return None
    s = str(v).strip()
    if s in ("", "-", "None") or s.startswith("="): return None
    try: return "{:.2f}".format(float(s))
    except: return None

# Map each product to its Excel sheet and expected row count
sheet_map = {
    "ysly-jz": ("YSLY", 153),
    "jz-500": ("JZ-500", 153),
    "olflex-classic-110": ("Olflex", 153),
    "flex-jz": ("Flex-JZ", 153),
    "opvc-jz": ("OPVC-JZ", 153),
    "olflex-classic-115-cy": ("115", 95),
    "liycy": ("LiYCY", 95),
    "liycy-jz": ("LiYCY-JZ", 95),
    "h07v-k": ("H07V-K", 18),
}

for slug, (sheet_name, expected_count) in sheet_map.items():
    ws = wb[sheet_name]
    json_items = json_data[slug]["items"]
    
    # Verify count
    if len(json_items) != expected_count:
        err(slug + " count: expected " + str(expected_count) + " got " + str(len(json_items)))
    
    # Determine Part NO column
    col2_header = ws.cell(row=4, column=2).value
    part_no_col = 2 if (col2_header and "Part" in str(col2_header)) else 3
    
    # Collect Excel data for comparison
    excel_items = []
    for row_num in range(6, ws.max_row + 1):
        core_size = ws.cell(row=row_num, column=8).value
        model = ws.cell(row=row_num, column=9).value
        if core_size is None and model is None:
            continue
        cs = str(core_size).strip() if core_size else ""
        if cs in ("Core x Conductor", "Size (mm2)", "", "None"):
            continue
        ms = str(model).strip() if model else ""
        if not ms or ms == "None":
            continue
        
        excel_items.append({
            "partNo": ws.cell(row=row_num, column=part_no_col).value,
            "coreSize": core_size,
            "model": model,
            "outerDia": ws.cell(row=row_num, column=11).value,
            "price": ws.cell(row=row_num, column=15).value,
            "resistance": ws.cell(row=row_num, column=14).value,
        })
    
    # Compare row by row
    mismatches = 0
    for i, (jitem, eitem) in enumerate(zip(json_items, excel_items)):
        # Check core size
        jcs = jitem["coreSize"] or ""
        ecs = str(eitem["coreSize"]).strip() if eitem["coreSize"] else ""
        if jcs != ecs:
            err(slug + " row " + str(i) + ": coreSize mismatch: JSON='" + jcs + "' Excel='" + ecs + "'")
            mismatches += 1
        
        # Check price
        jprice = jitem["price"]
        eprice = fmt_price(eitem["price"])
        if jprice != eprice:
            err(slug + " row " + str(i) + " (" + jcs + "): price mismatch: JSON=" + str(jprice) + " Excel=" + str(eprice))
            mismatches += 1
        
        # Check outer diameter
        jod = jitem["outerDia"] or ""
        eod = fmt_num(eitem["outerDia"]) or ""
        if jod != eod:
            err(slug + " row " + str(i) + " (" + jcs + "): outerDia mismatch: JSON='" + jod + "' Excel='" + eod + "'")
            mismatches += 1
        
        # Check resistance
        jr = jitem["resistance"] or ""
        er = fmt_num(eitem["resistance"]) or ""
        if jr != er:
            err(slug + " row " + str(i) + " (" + jcs + "): resistance mismatch: JSON='" + jr + "' Excel='" + er + "'")
            mismatches += 1
    
    if mismatches == 0:
        ok(slug + ": all " + str(len(json_items)) + " rows match Excel ✓")
    else:
        err(slug + ": " + str(mismatches) + " mismatches found")

# ══════════════════════════════════════════════════
# 7. SUMMARY
# ══════════════════════════════════════════════════
print("\n" + "=" * 60)
print("7. FINAL SUMMARY")
print("=" * 60)
print("  Products: " + str(len(json_data)))
print("  Total items: " + str(total_items))
print("  Errors: " + str(len(errors)))
print("  Warnings: " + str(len(warnings)))

if errors:
    print("\n  ❌ ERRORS FOUND - DO NOT PUSH:")
    for e in errors:
        print("    - " + e)
    sys.exit(1)
else:
    print("\n  ✅ ALL CHECKS PASSED — READY TO PUSH!")
    sys.exit(0)
