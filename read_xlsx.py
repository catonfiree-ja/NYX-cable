# -*- coding: utf-8 -*-
import openpyxl
import glob
import os

downloads = os.path.expanduser("~") + "\\Downloads"
files = [os.path.join(downloads, f) for f in os.listdir(downloads) if "NYX" in f and f.endswith(".xlsx")]
filepath = files[0]
print(f"File: {filepath}")
wb = openpyxl.load_workbook(filepath)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*60}")
    
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
        # Filter out completely empty rows
        if any(cell is not None for cell in row):
            cells = []
            for j, cell in enumerate(row):
                if cell is not None:
                    # Column labels
                    col_label = chr(65 + j)  # A, B, C, D...
                    val = str(cell).replace('\n', ' | ')
                    cells.append(f"  [{col_label}] {val}")
            if cells:
                print(f"Row {i}:")
                for c in cells:
                    print(c)
                print()
